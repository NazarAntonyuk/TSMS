from django.contrib import admin
from django.http import HttpResponse
from import_export.admin import ImportExportModelAdmin
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from .models import NonCurrentAsset
from .resources import NonCurrentAssetResource

import os
from django.conf import settings

@admin.register(NonCurrentAsset)
class NonCurrentAssetAdmin(ImportExportModelAdmin):
    resource_class = NonCurrentAssetResource
    list_display = [field.name for field in NonCurrentAsset._meta.fields if field.name != "id"]
    list_filter = [
        'journal_code', 'acquisition_date', 'placement',
        'decommissioning', 'person_in_charge'  # Add additional fields here for filtering
    ]
    search_fields = [field.name for field in NonCurrentAsset._meta.fields if field.name != "id"]
    actions = ["export_selected_xlsx"]

    def export_selected_xlsx(self, request, queryset):
        """
        Export selected objects to XLSX using a predefined Excel template, excluding the 'department_code' field.
        """
        resource = self.resource_class()
        dataset = resource.export(queryset)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Vybirka_asset.xlsx"'

        # Path to your Excel template
        template_path = os.path.join(settings.BASE_DIR, 'templates', 'non_current_asset', 'NonCurrentAsset_template.xlsx')

        # Load your template
        wb = load_workbook(filename=template_path)
        ws = wb.active

        # Start writing data from row 32
        data_start_row = 7
        headers = [header for header in dataset.headers if header != 'journal_code']
        for row_idx, data in enumerate(dataset.dict, start=data_start_row):
            for col_idx, header in enumerate(headers):
                cell = ws[f"{get_column_letter(col_idx + 1)}{row_idx}"]
                cell.value = data.get(header, '')

        # Save the modified file to the response
        wb.save(response)
        return response

    export_selected_xlsx.short_description = "Експортувати вибрані записи (XLSX)"

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        return form
