import os

from django.conf import settings
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import load_workbook

from .models import Auditorium, EquipmentInAuditorium


@admin.register(Auditorium)
class AuditoriumAdmin(admin.ModelAdmin):
    list_display = ('number', 'description_aud', 'head')
    search_fields = ('number', 'description_aud', 'head')

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.groups.filter(name='financially responsible KN').exists():
            # Вибір аудиторій, які може бачити користувач групи 'financially responsible KN'
            return qs.filter(number__in=['214', '213'])
        return qs


@admin.register(EquipmentInAuditorium)
class EquipmentInAuditoriumAdmin(admin.ModelAdmin):
    list_display = (
        'auditorium', 'equipment_details', 'get_inventory_number', 'get_quantity_asset', 'get_acquisition_date',
        'last_service_date', 'note_equip', 'repair_date'
    )
    list_filter = (
        'auditorium', 'last_service_date', 'repair_date'
    )
    search_fields = (
        'auditorium__number',
        'non_current_asset__inventory_number',
        'non_current_asset__name_description',
        'non_current_asset__factory_number',
        'non_current_asset__passport_number',
        'non_current_asset__unit_of_measurement',
        'non_current_asset__quantity_asset',
        'non_current_asset__initial_cost',
        'non_current_asset__arrival_mark',
        'non_current_asset__bux_initial_cost',
        'non_current_asset__bux_quantity_asset',
        'non_current_asset__wear_amount',
        'non_current_asset__balance_value',
        'non_current_asset__service_life',
        'non_current_asset__other_details',
        'non_current_asset__placement',
        'non_current_asset__decommissioning',
        'non_current_asset__person_in_charge',
        'equipment__processor',  # Припускаємо, що у Equipment є поле processor
        'equipment__ram_gb',
        'equipment__ram_type',
        'equipment__hdd_volume',
        'equipment__ssd_volume',
        'equipment__motherboard',
        'equipment__motherboard_socket',
        'equipment__power_supply',
        'equipment__operating_system',
        'equipment__office_software',
        'equipment__specialized_software',
        'equipment__screen_diagonal',
        'equipment__screen_resolution',
        'equipment__projector_type',
        'equipment__projector_resolution',
        'equipment__router_speed',
        'equipment__ports_number',
    )
    actions = ['delete_selected', 'export_selected_xlsx']

    def get_inventory_number(self, obj):
        return obj.non_current_asset.inventory_number if obj.non_current_asset else "—"

    get_inventory_number.short_description = 'Інвентарний номер'

    def get_name_description(self, obj):
        return obj.non_current_asset.name_description if obj.non_current_asset else "—"

    get_name_description.short_description = 'Найменування'

    def get_quantity_asset(self, obj):
        return obj.non_current_asset.quantity_asset if obj.non_current_asset else "—"

    get_quantity_asset.short_description = 'Кількість'

    def get_acquisition_date(self, obj):
        return obj.non_current_asset.acquisition_date if obj.non_current_asset else "—"

    get_acquisition_date.short_description = 'Дата придбання'

    def note_equip(self, obj):
        return obj.note_equip if hasattr(obj, 'note_equip') else "—"

    note_equip.short_description = 'Примітка'

    def equipment_details(self, obj):
        details = []
        if obj.equipment:
            for field in obj.equipment._meta.fields:
                value = getattr(obj.equipment, field.name, None)
                if value not in [None, '']:
                    details.append(f"{field.verbose_name}: {value}")
        return ' | '.join(details)

    equipment_details.short_description = 'Характеристики обладнання'

    def export_selected_xlsx(self, request, queryset):
        """
        Експорт вибраних об'єктів у XLSX з використанням шаблону.
        """
        # Шлях до вашого шаблону Excel
        template_path = os.path.join(settings.BASE_DIR, 'templates', 'auditorium',  'auditorium_templates.xlsx')

        # Завантаження шаблону робочої книги
        wb = load_workbook(template_path)
        ws = wb.active  # Припускаємо, що дані повинні бути записані в активний лист

        # Визначення початкового рядка для даних
        start_row = 2  # Відкоригуйте це значення в залежності від того, з якого рядка повинні починатися дані

        # Визначення заголовків, якщо вони ще не присутні в шаблоні
        headers = ['Аудиторія', 'Інвентарний номер', 'Найменування', 'Дата останнього сервісу', 'Дата ремонту',
                   'Примітка', 'Дата придбання', 'Характеристики обладнання']
        header_row = ws[1]
        if not all(cell.value for cell in header_row):
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_num, value=header)

        # Додавання рядків даних
        for row_num, obj in enumerate(queryset, start=start_row):
            equipment_details = []
            if obj.equipment:
                for field in obj.equipment._meta.fields:
                    value = getattr(obj.equipment, field.name, None)
                    if value not in [None, '']:
                        equipment_details.append(f"{field.verbose_name}: {value}")

            row = [
                obj.auditorium.number if obj.auditorium else "",
                obj.non_current_asset.inventory_number if obj.non_current_asset else "",
                obj.non_current_asset.name_description if obj.non_current_asset else "",
                obj.non_current_asset.quantity_asset if obj.non_current_asset else "",
                obj.last_service_date.strftime("%Y-%m-%d") if obj.last_service_date else "",
                obj.repair_date.strftime("%Y-%m-%d") if obj.repair_date else "",
                obj.note_equip,
                obj.non_current_asset.acquisition_date if obj.non_current_asset and obj.non_current_asset.quantity_asset and obj.non_current_asset.acquisition_date else "",
                ' | '.join(equipment_details)
            ]
            for col_num, value in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Збереження робочої книги у відповідь
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="vybirka_equipment.xlsx"'
        wb.save(response)
        return response

    export_selected_xlsx.short_description = "Експортувати вибрані (XLSX)"

    def get_form(self, request, obj=None, **kwargs):
        form = super().get_form(request, obj, **kwargs)
        return form
